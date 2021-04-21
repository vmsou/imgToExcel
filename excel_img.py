import sys
import cv2 as cv
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Color
from openpyxl.utils.cell import get_column_letter
import numpy as np
from time import perf_counter


def evaluate_speed():
    widt, heig = 5, 5
    eval_start = perf_counter()
    temp_wb = openpyxl.Workbook()
    fake_sheet = temp_wb.active
    fake_img = np.zeros((widt, heig, 3), dtype='uint8')

    for column in range(1, widt + 1):
        for row in range(1, heig + 1):
            pos = ((column - 1) * 3) + 1
            fake_img[column - 1][row - 1][0] = np.random.randint(0, 255)
            fake_img[column - 1][row - 1][1] = np.random.randint(0, 255)
            fake_img[column - 1][row - 1][2] = np.random.randint(0, 255)
            fake_sheet.cell(row=row, column=pos).value = int(fake_img[column - 1][row - 1][0])
            fake_sheet.cell(row=row, column=pos + 1).value = int(fake_img[column - 1][row - 1][1])
            fake_sheet.cell(row=row, column=pos + 2).value = int(fake_img[column - 1][row - 1][2])

    temp_wb.close()
    pixels = widt * heig
    eval_end = perf_counter()

    pix_per_second = pixels / (eval_end - eval_start)

    return pix_per_second * 2

pxl_per_sec = evaluate_speed()  # 10146.3314070309

red = [Color('000000'), Color('ff0000')]
green = [Color('000000'), Color('00ff00')]
blue = [Color('000000'), Color('0000ff')]

red_col = ColorScaleRule(start_type='num', start_value=0, start_color=red[0],
                         end_type='num', end_value=255, end_color=red[1])
green_col = ColorScaleRule(start_type='num', start_value=0, start_color=green[0],
                           end_type='num', end_value=255, end_color=green[1])
blue_col = ColorScaleRule(start_type='num', start_value=0, start_color=blue[0],
                          end_type='num', end_value=255, end_color=blue[1])


def adjust_columns(sheet, rgb_colors, width_size):
    for column in range(1, len(rgb_colors[0])):
        pos = ((column - 1) * 3) + 1
        to_r = get_column_letter(pos + 2)  # pos
        to_g = get_column_letter(pos + 1)  # pos + 1
        to_b = get_column_letter(pos)  # pos + 2
        sheet.conditional_formatting.add(f'{to_r}1:{to_r}{len(rgb_colors)}', red_col)
        sheet.conditional_formatting.add(f'{to_g}1:{to_g}{len(rgb_colors)}', green_col)
        sheet.conditional_formatting.add(f'{to_b}1:{to_b}{len(rgb_colors)}', blue_col)
        sheet.column_dimensions[to_r].width = width_size
        sheet.column_dimensions[to_g].width = width_size
        sheet.column_dimensions[to_b].width = width_size


def adjust_rows(sheet, rgb_colors, height_size):
    for row in range(1, len(rgb_colors) + 1):
        sheet.row_dimensions[row].height = height_size


def img_to_excel(sheet, rgb_colors):
    for column in range(1, len(rgb_colors[0])):
        for row in range(1, len(rgb_colors)):
            rgb_pos = ((column - 1) * 3) + 1

            try:
                # red
                sheet.cell(row=row, column=rgb_pos + 2).value = int(rgb_colors[row - 1][column - 1][2])  # rgbpos and 0
                # green
                sheet.cell(row=row, column=rgb_pos + 1).value = int(rgb_colors[row - 1][column - 1][1])  # rgbpos  + 1 and 1
                # blue
                sheet.cell(row=row, column=rgb_pos).value = int(rgb_colors[row - 1][column - 1][0])  # # rgbpos + 2 and 2
            except IndexError as e:
                print(f"column: {column}, {rgb_pos}, row: {row}, {e}")


RATIO = 1

def main():
    global RATIO
    IMG_PATH = input("Image Path: ")
    colors = cv.imread(IMG_PATH)
    hg, wid, _ = colors.shape
    print(f"For an image {wid} x {hg} it will take aprox. {((hg / RATIO)  * (wid / RATIO)) / pxl_per_sec:.2f}"
          f" seconds at RATIO: {RATIO}")
    proceed = input('Do you want to continue or change the ratio? (continue/change):')
    end_proceed = False
    if proceed.lower() == "change":
        while not end_proceed:
            RATIO = int(input("RATIO: "))
            print(f"For an image {wid} x {hg} it will take aprox. {((hg / RATIO) * (wid / RATIO)) / pxl_per_sec:.2f} seconds at RATIO: {RATIO}")
            confirm = input("Confirm (y/n/exit): ")
            if confirm.lower() == 'y':
                end_proceed = True
            elif confirm.lower() == 'exit':
                sys.exit()

    start = perf_counter()

    wid, hg = int(wid / RATIO), int(hg / RATIO)

    rgb_colors = cv.resize(colors, (wid, hg), cv.INTER_AREA)
    #rgb_colors = cv.cvtColor(colors, cv.COLOR_BGR2RGB)  # not converting to speed up time, but if you want to, adjust column values

    width_size = 1
    height_size = 20

    wb = openpyxl.Workbook()
    sheet = wb.active

    adjust_columns(sheet, rgb_colors, width_size)
    adjust_rows(sheet, rgb_colors, height_size)
    img_to_excel(sheet, rgb_colors)

    wb.save(f"{IMG_PATH.split('.')[0]}.xlsx")
    end = perf_counter()

    print(f'It took {end - start:.2f} seconds')


if __name__ == '__main__':
    main()
